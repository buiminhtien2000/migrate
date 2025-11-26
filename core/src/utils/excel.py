# excel.py
import os
from typing import Dict, Optional, List, Any
import pandas as pd # type: ignore
from openpyxl.styles import PatternFill, Font # type: ignore
from openpyxl.utils import get_column_letter # type: ignore
import aiohttp # type: ignore
import asyncio
from io import BytesIO
import re
from concurrent.futures import ThreadPoolExecutor

class Excel:
    def __init__(self, logger: 'CustomLogger', output_path: Optional[str] = None): # type: ignore
        self.output_path = output_path
        self.logger = logger
        self.session = None

    async def _save_formatted_excel(self, df: pd.DataFrame, filename: str) -> None:
        """Save DataFrame to Excel with formatting."""
        try:
            # Đảm bảo thư mục tồn tại
            os.makedirs(os.path.dirname(filename), exist_ok=True)
            
            with pd.ExcelWriter(filename, engine='openpyxl', mode='w') as writer:
                # Làm sạch tên cột
                df.columns = [str(col).strip().replace('[', '').replace(']', '') for col in df.columns]
                
                df.to_excel(writer, index=False, sheet_name='Failed Items')
                worksheet = writer.sheets['Failed Items']
                
                # Format headers
                for col_num, value in enumerate(df.columns.values):
                    cell = worksheet.cell(1, col_num + 1)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color='D3D3D3',
                        end_color='D3D3D3',
                        fill_type='solid'
                    )

                # Auto-adjust column widths
                for idx, col in enumerate(df.columns):
                    try:
                        max_length = max(
                            df[col].astype(str).apply(len).max(),
                            len(str(col))
                        )
                        col_letter = get_column_letter(idx + 1)
                        worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)
                    except Exception as e:
                        self.logger.warning(f"Could not set column width for {col}: {str(e)}")
                        continue

        except Exception as e:
            self.logger.error(f"Error formatting Excel file: {str(e)}", exc_info=True)
            raise

    def _convert_drive_url(self, url: str) -> str:
        """
        Chuyển đổi Google Drive/Sheets sharing URL sang direct download URL
        
        Supports:
        - https://drive.google.com/file/d/{file_id}/view?usp=sharing
        - https://drive.google.com/open?id={file_id}
        - https://docs.google.com/spreadsheets/d/{file_id}/edit?usp=sharing
        """
        try:
            # Pattern cho file_id
            patterns = [
                r'\/file\/d\/([a-zA-Z0-9-_]+)',  # Format: /file/d/{file_id}
                r'[?&]id=([a-zA-Z0-9-_]+)',      # Format: ?id={file_id} or &id={file_id}
                r'\/spreadsheets\/d\/([a-zA-Z0-9-_]+)', # Format: /spreadsheets/d/{file_id}
            ]
            
            file_id = None
            for pattern in patterns:
                match = re.search(pattern, url)
                if match:
                    file_id = match.group(1)
                    break
            
            if not file_id:
                raise ValueError("Invalid Google Drive/Sheets URL format")
            
            # Nếu là Google Sheets
            if 'docs.google.com/spreadsheets' in url:
                return f'https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx'
            
            # Nếu là Google Drive
            return f'https://drive.google.com/uc?export=download&id={file_id}'
            
        except Exception as e:
            self.logger.error(f"Error converting Google URL: {str(e)}")
            return url

    def export_to_file(self, data: list, filename: str) -> None:
        """
        Export data to Excel file with formatting
        
        Args:
            data: List of dictionaries to export
            filename: Output Excel file path
            
        Features:
            - Auto column width based on content
            - Gray header with bold text
            - Clean formatting
        """
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Write to Excel
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        df.to_excel(writer, index=False)
        
        # Get worksheet for formatting
        worksheet = writer.sheets['Sheet1']
        
        # Format header row
        header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        header_font = Font(bold=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Auto adjust column width based on content
        for idx, column in enumerate(df.columns):
            column_letter = get_column_letter(idx + 1)
            max_length = max(
                df[column].astype(str).apply(len).max(),  # Max length of data
                len(str(column))  # Length of header
            )
            adjusted_width = min(max_length + 2, 50)  # Add padding, max 50
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save file
        writer.close()
              
    async def save_failed_items(self, failed_item: Dict, filename: str) -> None:
        """Save failed item to Excel file with formatting."""
        try:
            # Convert single item to DataFrame
            new_df = pd.DataFrame([failed_item])
                
            if os.path.exists(filename):
                # Thêm engine='openpyxl' khi đọc file
                existing_df = pd.read_excel(filename, engine='openpyxl')
                df = pd.concat([existing_df, new_df], ignore_index=True)
            else:
                df = new_df

            # Optimize columns order
            important_cols = ['timestamp', 'processor', 'error', 'status', 'item']
            other_cols = [col for col in df.columns if col not in important_cols]
            existing_important_cols = [col for col in important_cols if col in df.columns]
            df = df[existing_important_cols + other_cols]

            await self._save_formatted_excel(df, filename)
            
        except Exception as e:
            self.logger.error(f"Error saving failed items: {str(e)}", exc_info=True)
            raise

    def get_mappings_sync(self, url: str) -> Dict[str, Dict[str, str]]:
        """
        Sync wrapper for getting Excel mappings
        """
        try:
            with ThreadPoolExecutor(max_workers=1) as executor:
                def run_async():
                    loop = asyncio.new_event_loop()
                    asyncio.set_event_loop(loop)
                    try:
                        if self.session is None or self.session.closed:
                            self.session = aiohttp.ClientSession()
                        return loop.run_until_complete(self.get_field_mappings_from_url(url))
                    finally:
                        if self.session and not self.session.closed:
                            loop.run_until_complete(self.session.close())
                        loop.close()

                future = executor.submit(run_async)
                return future.result()
        except Exception as e:
            self.logger.error(f"Error getting mappings: {e}")
            return {}

    async def get_field_mappings_from_url(self, url: str) -> Dict[str, Dict[str, str]]:
        """
        Đọc file Excel mapping từ URL và trả về dictionary mapping fields.
        Session được handle tự động trong hàm.
        """
        try: 
            # Chuyển đổi URL nếu là Google Drive sharing link
            if 'drive.google.com' in url or 'docs.google.com' in url:
                download_url = self._convert_drive_url(url)
            else:
                download_url = url

            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
            
            async with self.session.get(download_url, headers=headers, timeout=30) as response:
                if response.status != 200:
                    raise aiohttp.ClientResponseError(
                        response.request_info,
                        response.history,
                        status=response.status
                    )

                # Kiểm tra nếu Google Drive yêu cầu xác nhận
                if 'confirm=' in str(response.url):
                    confirm_token = str(response.url).split('confirm=')[1]
                    download_url = f"{download_url}&confirm={confirm_token}"
                    async with self.session.get(download_url, headers=headers, timeout=30) as confirmed_response:
                        if confirmed_response.status != 200:
                            raise aiohttp.ClientResponseError(
                                confirmed_response.request_info,
                                confirmed_response.history,
                                status=confirmed_response.status
                            )
                        content = await confirmed_response.read()
                else:
                    content = await response.read()

            # Đọc Excel từ content
            excel_data = BytesIO(content)
            excel_file = pd.ExcelFile(excel_data)
            
            # Dictionary để lưu kết quả
            mappings = {}
            
            # Đọc từng sheet
            for sheet_name in excel_file.sheet_names:
                try:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name, usecols=[0,1])
                    
                    source_col = df.columns[0]
                    target_col = df.columns[1]
                    
                    df = df.dropna(how='all')
                    df[source_col] = df[source_col].astype(str).str.strip()
                    df[target_col] = df[target_col].astype(str).str.strip()
                    
                    sheet_mapping = dict(zip(df[source_col], df[target_col]))
                    
                    # Chuẩn hóa format key
                    fkey = sheet_name.strip().lower()
                    mappings[fkey] = sheet_mapping
                    
                except Exception as sheet_error:
                    self.logger.error(f"Error processing sheet {sheet_name}: {str(sheet_error)}")
                    continue
                    
            self.mappings = mappings
            return mappings
            
        except aiohttp.ClientError as client_error:
            self.logger.error(f"Network error downloading file: {str(client_error)}")
            return {}
        except Exception as e:
            self.logger.error(f"Error processing Excel file: {str(e)}")
            return {}

    def export_to_spreadsheet(self, filename: str, data_list: List[Dict[str, Any]]) -> str:
        """
        Convert multiple data sources into a single Excel file with multiple sheets
        
        Args:
            filename (str): Name of the output file (e.g: "output.xlsx") 
            data_list (List[Dict]): List of dictionaries containing:
                - sheet_name: str - Name for the sheet
                - data: Union[List, Dict, pd.DataFrame] - Data for the sheet
        
        Returns:
            str: Filename of the created Excel file
        """
        # Create Excel writer object
        file_path = f"{self.output_path}/{filename}"
        writer = pd.ExcelWriter(file_path, engine='xlsxwriter')
        workbook = writer.book
        
        # Create header format
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#D3D3D3',
            'border': 1,
            'text_wrap': True,
            'valign': 'vcenter',
            'align': 'center'
        })
        
        # Process each data source
        for data in data_list:
            try:
                raw_data = data['data']
                sheet_name = data['name'][:31]  # Excel has 31 char limit
                
                # Convert data to DataFrame based on type
                if isinstance(raw_data, pd.DataFrame):
                    df = raw_data
                elif isinstance(raw_data, list):
                    df = pd.DataFrame(raw_data)
                elif isinstance(raw_data, dict):
                    df = pd.DataFrame([raw_data])
                else:
                    raise ValueError(f"Unsupported data type for sheet {sheet_name}")
                
                # Write DataFrame to Excel sheet
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Get worksheet object
                worksheet = writer.sheets[sheet_name]
                
                # Format the worksheet
                for col_num, value in enumerate(df.columns.values):
                    worksheet.write(0, col_num, value, header_format)
                    # Set column width based on maximum length
                    max_length = max(
                        df[value].astype(str).apply(len).max(),  # max length of values
                        len(str(value))  # length of header
                    )
                    worksheet.set_column(col_num, col_num, min(max_length + 2, 30))
                
                # Add autofilter
                worksheet.autofilter(0, 0, len(df), len(df.columns)-1)
                
                # Freeze top row
                worksheet.freeze_panes(1, 0)
                
            except Exception as e:
                self.logger.error(f"Error processing sheet {sheet_name}: {str(e)}")
                continue
        
        # Save and close the writer
        writer.close()
        
        return filename
    
    async def close(self):
        """Close aiohttp session"""
        if self.session and not self.session.closed:
            await self.session.close()      